"""
RPA Module: Tracker Updater
Handles all writes to the applications_master.xlsx Excel file.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path
from datetime import datetime
import pandas as pd

from config.settings import TRACKER_PATH

def initialise_tracker():
    """Creates the master Excel tracker with headers if it doesn't exist."""
    if TRACKER_PATH.exists():
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    headers = [
        "ref_number", "applicant_name", "applicant_email", "received_at",
        "status", "doc_check_passed", "missing_docs", "income_declared",
        "income_payslip", "income_ratio", "income_score", "income_discrepancy",
        "employer_ref_score", "landlord_ref_score", "cover_letter_score",
        "inconsistency_risk", "composite_score", "recommendation",
        "invite_to_viewing", "action_required", "gdpr_consent",
        "gdpr_deletion_due", "acknowledgement_sent", "rejection_sent",
        "missing_doc_email_sent", "notes"
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    wb.save(TRACKER_PATH)
    print(f"[TRACKER] Initialised at {TRACKER_PATH}")

def add_application_row(app: dict):
    """Adds a new application row to the tracker."""
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active
    
    # Check if already exists
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == app["ref_number"]:
            return

    new_row = [app.get("ref_number"), app.get("applicant_name"), app.get("applicant_email"), 
               app.get("received_at"), app.get("status"), "", "", "", "", "", "", "", "", "", 
               "", "", "", "", "", "", app.get("consent_given"), "", "", "", "", ""]
    ws.append(new_row)
    wb.save(TRACKER_PATH)

def update_application_status(ref_number: str, status: str):
    """Updates the status column for a given application."""
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == ref_number:
            row[4].value = status
            break
    wb.save(TRACKER_PATH)

def update_document_flags(ref_number: str, doc_status: dict):
    """Updates doc_check_passed and missing_docs columns."""
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active
    missing = [k for k, v in doc_status.items() if v != "PRESENT"]
    passed = len(missing) == 0
    for row in ws.iter_rows(min_row=2):
        if row[0].value == ref_number:
            row[5].value = passed
            row[6].value = ", ".join(missing)
            break
    wb.save(TRACKER_PATH)

def update_income_data(app: dict):
    """Updates income-related columns."""
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == app["ref_number"]:
            row[7].value = app.get("declared_income")
            row[8].value = app.get("authoritative_income")
            row[9].value = app.get("income_ratio")
            row[10].value = app.get("income_score")
            row[11].value = str(app.get("income_discrepancy")) if app.get("income_discrepancy") else ""
            row[4].value = app.get("status")
            break
    wb.save(TRACKER_PATH)

def get_existing_refs() -> set:
    """Returns a set of all reference numbers already in the tracker."""
    if not TRACKER_PATH.exists():
        return set()
    try:
        df = pd.read_excel(TRACKER_PATH)
        if df.empty:
            return set()
        return set(df["ref_number"].astype(str).tolist())
    except Exception:
        return set()

def finalise_tracker(applications: list[dict]):
    """Final bulk update for any remaining fields."""
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active
    app_map = {a["ref_number"]: a for a in applications}
    
    for row in ws.iter_rows(min_row=2):
        ref = row[0].value
        if ref in app_map:
            app = app_map[ref]
            row[4].value = app.get("status", row[4].value)
            row[12].value = app.get("employer_ref_score")
            row[13].value = app.get("landlord_ref_score")
            row[14].value = app.get("cover_letter_score")
            row[15].value = app.get("inconsistency_risk")
            row[16].value = app.get("composite_score")
            row[17].value = app.get("recommendation")
            row[18].value = app.get("invite_to_viewing")
            row[19].value = app.get("action_required")
            
    wb.save(TRACKER_PATH)
